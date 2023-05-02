import tkinter
import customtkinter as ctk
# from tkinter import ttk
# from tkinter import messagebox
import os
import openpyxl
from tkcalendar import Calendar, DateEntry
from PIL import ImageTk,Image
from docxtpl import DocxTemplate
import datetime
import docx2pdf

def enter_data():
    #accepted = accept_var.get()
    
        # User info
    name = entry_name.get()
    surname = entry_surname.get()
    kw = entry_kw.get()
        
    if name and surname:
            year = combobox_yea_r.get()
            # age = age_spinbox.get()
            course = combobox_cours_e.get()
            date_from = calender_datefrom.get_date()
            date_to = calender_dateto.get_date()

            # Course info
            # registration_status = reg_status_var.get()
            duration_monday = entry_duration_monday.get()
            duration_tuesday = entry_duration_tuesday.get()
            duration_wednesday = entry_duration_wednesday.get()
            duration_thursday = entry_duration_thursday.get()
            duration_friday = entry_duration_friday.get()

            content_monday = entry_content_monday.get(0.0, "end")
            content_tuesday = entry_content_tuesday.get(0.0, "end")
            content_wednesday = entry_content_wendesday.get(0.0, "end")
            content_thursday = entry_content_thursday.get(0.0, "end")
            content_friday = entry_content_friday.get(0.0, "end")
            
            print("Name: ", name, "Surame: ", surname, "KW: ", kw)
            print("Course: ", course, "Year: ", year, "Date From: ", date_from, "Date To: ", date_to)
            print("# Duration: ", duration_monday, "# Content: ", content_monday)
            print("# Duration: ", duration_tuesday, "# Content: ", content_tuesday)
            print("# Duration: ", duration_wednesday, "# Content: ", content_wednesday)
            print("# Duration: ", duration_thursday, "# Content: ", content_thursday)
            print("# Duration: ", duration_friday, "# Content: ", content_friday)
            print("------------------------------------------")
            
            excel_template = f"template\excel-template.xlsx"
            filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.xlsx"
            
            if not os.path.exists(filepath):
                file = openpyxl.load_workbook(excel_template)
                sheet = file.active
                #----------------------------------------------------------------
                sheet['A2'].value = name
                sheet['B2'].value = surname
                sheet['C2'].value = kw
                sheet['D2'].value = year
                sheet['E2'].value = course
                sheet['F2'].value = date_from
                sheet['G2'].value = date_to
                #----------------------------------------------------------------
                sheet['A5'].value = duration_monday
                sheet['B5'].value = duration_tuesday
                sheet['C5'].value = duration_wednesday
                sheet['D5'].value = duration_thursday
                sheet['E5'].value = duration_friday
                #----------------------------------------------------------------
                sheet['B8'].value = content_monday
                sheet['B11'].value = content_tuesday
                sheet['B14'].value = content_wednesday
                sheet['B17'].value = content_thursday
                sheet['B20'].value = content_friday
            #----------------------------------------------------------------

            # heading = ["Name", "Surname", "KW", "Course", "Year", "Date From", "Date To", "Duration Monday", "Duration Tuesday", "Duration Wednesday", "Duration Thursday", "Duration friday", "Content Monday", "Content Tuesday", "Content Wednesday", "Content Thursday", "Content friday"]
            # sheet.append(heading)
            # workbook.save(filepath)
            # workbook = openpyxl.load_workbook(filepath)
            # sheet = workbook.active
            # sheet.append([name, surname, kw, course, year, date_from, date_to, duration_monday, duration_tuesday, duration_wednesday, duration_thursday, duration_friday, content_monday, content_tuesday, content_wednesday, content_thursday, content_friday
            #               ])
                filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.xlsx"
                file.save(filepath)
                tkinter.messagebox.showinfo(title="Success", message="The data was saved successfully and the file was created successfully.")
            
            if os.path.exists(filepath):
                file = openpyxl.load_workbook(filepath)
                sheet = file.active
                #----------------------------------------------------------------
                sheet['A2'].value = name
                sheet['B2'].value = surname
                sheet['C2'].value = kw
                sheet['E2'].value = year
                sheet['D2'].value = course
                sheet['F2'].value = date_from
                sheet['G2'].value = date_to
                #----------------------------------------------------------------
                sheet['A5'].value = duration_monday
                sheet['B5'].value = duration_tuesday
                sheet['C5'].value = duration_wednesday
                sheet['D5'].value = duration_thursday
                sheet['E5'].value = duration_friday
                #----------------------------------------------------------------
                sheet['B8'].value = content_monday
                sheet['B11'].value = content_tuesday
                sheet['B14'].value = content_wednesday
                sheet['B17'].value = content_thursday
                sheet['B20'].value = content_friday
            #----------------------------------------------------------------

            # heading = ["Name", "Surname", "KW", "Course", "Year", "Date From", "Date To", "Duration Monday", "Duration Tuesday", "Duration Wednesday", "Duration Thursday", "Duration friday", "Content Monday", "Content Tuesday", "Content Wednesday", "Content Thursday", "Content friday"]
            # sheet.append(heading)
            # workbook.save(filepath)
            # workbook = openpyxl.load_workbook(filepath)
            # sheet = workbook.active
            # sheet.append([name, surname, kw, course, year, date_from, date_to, duration_monday, duration_tuesday, duration_wednesday, duration_thursday, duration_friday, content_monday, content_tuesday, content_wednesday, content_thursday, content_friday
            #               ])
                filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.xlsx"
                file.save(filepath)
                tkinter.messagebox.showinfo(title="Success", message="The data was saved successfully and the file was created successfully.")


                
    else:
            tkinter.messagebox.showwarning(title="Error", message="Name, Surname, KW and Year are required.")
    # else:
    #     tkinter.messagebox.showwarning(title= "Error", message="You have not accepted the terms")


def load_file():

    name = entry_name.get()
    surname = entry_surname.get()
    kw = entry_kw.get()
    year = combobox_yea_r.get()

    filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.xlsx"
        
    if name and surname and kw:
        if os.path.exists(filepath):
            file = openpyxl.load_workbook(filepath)
            sheet = file.active


            # name = sheet['A2'].value
            # surname = sheet['B2'].value
            # kw = sheet['C2'].value
            # year = sheet['E2'].value
            # age = age_spinbox.get()
            course = sheet['D2'].value
            date_from = sheet['F2'].value
            date_to = sheet['G2'].value

            # Course info
            # registration_status = reg_status_var.get()
            duration_monday = sheet['A5'].value
            duration_tuesday = sheet['B5'].value
            duration_wednesday = sheet['C5'].value
            duration_thursday = sheet['D5'].value
            duration_friday = sheet['E5'].value

            content_monday = sheet['B8'].value
            content_tuesday = sheet['B11'].value
            content_wednesday = sheet['B14'].value
            content_thursday = sheet['B17'].value
            content_friday = sheet['B20'].value


            # combobox_cours_e.insert(0,f'{course}')
            entry_duration_monday.insert(0, f'{duration_monday}')
            entry_duration_tuesday.insert(0 ,f'{duration_tuesday}')
            entry_duration_wednesday.insert(0 ,f'{duration_wednesday}')
            entry_duration_thursday.insert(0 ,f'{duration_thursday}')
            entry_duration_friday.insert(0 ,f'{duration_friday}')
            entry_content_monday.insert("0.0",f'{content_monday}')
            entry_content_tuesday.insert("0.0",f'{content_tuesday}')
            entry_content_wendesday.insert("0.0",f'{content_wednesday}')
            entry_content_thursday.insert("0.0",f'{content_thursday}')
            entry_content_friday.insert("0.0",f'{content_friday}')
            
            filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.xlsx"
            tkinter.messagebox.showinfo(title="Success", message="The data has been successfully loaded.")

        else:
            tkinter.messagebox.showwarning(title="Error", message="File does not exist.")  

    else:
        tkinter.messagebox.showwarning(title="Error", message="Name, Surname, KW and Year are required.")  
    
def clear():
        entry_name.delete(0, "end")
        entry_surname.delete(0, "end")
        entry_kw.delete(0, "end")
        entry_duration_monday.delete("0", "end")
        entry_duration_tuesday.delete("0", "end")
        entry_duration_wednesday.delete("0", "end")
        entry_duration_thursday.delete("0", "end")
        entry_duration_friday.delete("0", "end")
        entry_content_monday.delete("0.0", "end")
        entry_content_tuesday.delete("0.0", "end")
        entry_content_wendesday.delete("0.0", "end")
        entry_content_thursday.delete("0.0", "end")
        entry_content_friday.delete("0.0", "end")

def word_to_pdf():
     
    name = entry_name.get()
    surname = entry_surname.get()
    kw = entry_kw.get()
    year = combobox_yea_r.get()

    pdf_filepath_source = f"Exports_Word\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.docx"
    pdf_filepath_output = f"Exports_PDF\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.pdf"
        
    if name and surname and kw:
        if os.path.exists(pdf_filepath_source):
             
            docx2pdf.convert(pdf_filepath_source, pdf_filepath_output)
            tkinter.messagebox.showinfo(title="Success", message="Report has been saved successfully into a PDF file.")
        
        else:
            tkinter.messagebox.showwarning(title="Error", message="File not found. Please correct the inputs and try again.")
    
    else:
        tkinter.messagebox.showwarning(title="Error", message="Name, Surname, KW and Year are required.")   


def excel_to_word():
     
    name = entry_name.get()
    surname = entry_surname.get()
    kw = entry_kw.get()
    year = combobox_yea_r.get()
    course = combobox_cours_e.get()
    date_from = calender_datefrom.get_date()
    date_to = calender_dateto.get_date()

            # Course info
            # registration_status = reg_status_var.get()
    duration_monday = entry_duration_monday.get()
    duration_tuesday = entry_duration_tuesday.get()
    duration_wednesday = entry_duration_wednesday.get()
    duration_thursday = entry_duration_thursday.get()
    duration_friday = entry_duration_friday.get()

    content_monday = entry_content_monday.get(0.0, "end")
    content_tuesday = entry_content_tuesday.get(0.0, "end")
    content_wednesday = entry_content_wendesday.get(0.0, "end")
    content_thursday = entry_content_thursday.get(0.0, "end")
    content_friday = entry_content_friday.get(0.0, "end")

    filepath = f"Exports_Excel\Arbeitsbericht_KW_{kw}_{year}_{name}_{surname}_.xlsx"
        
    if name and surname and kw and year:
        if os.path.exists(filepath):
            file = openpyxl.load_workbook(filepath)
            sheet = file.active


            name = sheet['A2'].value
            surname = sheet['B2'].value
            kw = sheet['C2'].value
            year = sheet['E2'].value
            # age = age_spinbox.get()
            course = sheet['D2'].value
            date_from = sheet['F2'].value
            date_to = sheet['G2'].value

            # Course info
            # registration_status = reg_status_var.get()
            duration_monday = sheet['A5'].value
            duration_tuesday = sheet['B5'].value
            duration_wednesday = sheet['C5'].value
            duration_thursday = sheet['D5'].value
            duration_friday = sheet['E5'].value

            content_monday = sheet['B8'].value
            content_tuesday = sheet['B11'].value
            content_wednesday = sheet['B14'].value
            content_thursday = sheet['B17'].value
            content_friday = sheet['B20'].value
            # Generate docs
            doc = DocxTemplate("template\example.docx")

            doc.render({    "name": name,
                            "surname": surname,
                            "kw": kw,
                            "year": year,
                            "course": course,
                            "date_from": date_from,
                            "date_to": date_to,
                            "duration_monday": duration_monday,
                            "duration_tuesday": duration_tuesday,
                            "duration_wednesday": duration_wednesday,
                            "duration_thursday": duration_thursday,
                            "duration_friday": duration_friday,
                            "content_monday": content_monday,
                            "content_tuesday": content_tuesday,
                            "content_wednesday": content_wednesday,
                            "content_thursday": content_thursday,
                            "content_friday": content_friday,
                            })
                
            doc_name = "Exports_Word\Arbeitsbericht_KW_" + kw + "_" + year + "_" + name + "_" + surname + "_.docx"
                
            doc.save(doc_name)
            tkinter.messagebox.showinfo(title="Success", message="Report has been saved successfully into a Word file.")
    
        else:
            tkinter.messagebox.showwarning(title="Error", message="File not found. Please correct the inputs and try again.")
    
    else:
        tkinter.messagebox.showwarning(title="Error", message="Name, Surname, KW and Year are required.")

# window = tkinter.Tk()
# window.title("Data Entry Form")

# frame = tkinter.Frame(window)
# frame.pack()

app = ctk.CTk()
app.title("XReports APP")
app.iconbitmap("XReports_APP_icon.ico")

def close():
     app.destroy()

# Saving User Info

img=ImageTk.PhotoImage(Image.open("UI_1.png").resize((1200,663)))
l1=ctk.CTkLabel(master=app, text="", font=('Century Gothic',20), image=img)
l1.pack()

#------------------------------------------------------------------------------------------------------------------------

# ----------------------------------------> User Information <---------------------------------------------------------

#------------------------------------------------------------------------------------------------------------------------


label_user_information=ctk.CTkLabel(master=l1, text="User information", bg_color='#ACACAC', font=('Century Gothic',12))
label_user_information.place(x=41, y=47)

#---------------------------------------------------------------------------------------------------------

#---------------------------------------------------------------------------------------------------------

label_kw=ctk.CTkLabel(master=l1, text="KW", bg_color='#ACACAC', font=('Century Gothic',12))
label_kw.place(x=306, y=30)

entry_kw=ctk.CTkEntry(master=l1, width=86, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_kw.place(x=280, y=50)

#---------------------------------------------------------------------------------------------------------

label_name=ctk.CTkLabel(master=l1, text="Name", bg_color='#ACACAC', font=('Century Gothic',12))
label_name.place(x=83, y=84)

entry_name=ctk.CTkEntry(master=l1, width=86, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_name.place(x=59, y=111)

#---------------------------------------------------------------------------------------------------------

label_surname=ctk.CTkLabel(master=l1, text="Surname", bg_color='#ACACAC', font=('Century Gothic',12))
label_surname.place(x=186, y=84)

entry_surname=ctk.CTkEntry(master=l1, width=86, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_surname.place(x=170, y=111)

#---------------------------------------------------------------------------------------------------------

label_year=ctk.CTkLabel(master=l1, text="Year", bg_color='#ACACAC', font=('Century Gothic',12))
label_year.place(x=306, y=84)

def combobox_year(choice):
    print("combobox dropdown clicked:", choice)

combobox_yea_r = ctk.CTkComboBox(master=l1,
                                    values=["2023", "2024", "2025", "2026", "2027", "2028", "2029", "2030", "2031", "2032", "2033", "2034"],
                                     command=combobox_year,
                                     bg_color='#ACACAC',
                                     width=86)
combobox_yea_r.place(x=281, y=111)
combobox_yea_r.set("2023")  # set initial value

#---------------------------------------------------------------------------------------------------------

label_course=ctk.CTkLabel(master=l1, text="Course", bg_color='#ACACAC', font=('Century Gothic',12))
label_course.place(x=83, y=156)

def combobox_course(choice):
    print("combobox dropdown clicked:", choice)

combobox_cours_e = ctk.CTkComboBox(master=l1,
                                     values=["", "Fachinformatiker für Anwendungsentwicklung", "Fachinformatiker für Systemintegration"],
                                     command=combobox_course,
                                     bg_color='#ACACAC',
                                     width=312)
combobox_cours_e.place(x=59, y=181)
combobox_cours_e.set("Fachinformatiker für Systemintegration")  # set initial value

#------------------------------------------------------------------------------------------------------------------------

# ----------------------------------------> Report Information <---------------------------------------------------------

#------------------------------------------------------------------------------------------------------------------------


label_report_information=ctk.CTkLabel(master=l1, text="Report information", bg_color='#ACACAC', font=('Century Gothic',12))
label_report_information.place(x=41, y=255)

label_datefrom=ctk.CTkLabel(master=l1, text="Date from", bg_color='#ACACAC', font=('Century Gothic',12))
label_datefrom.place(x=73, y=292)

calender_datefrom = DateEntry(master=l1, width=8, background= "#000000", bd=5)
calender_datefrom.place(x=59, y=319)

label_dateto=ctk.CTkLabel(master=l1, text="Date to", bg_color='#ACACAC', font=('Century Gothic',12))
label_dateto.place(x=190, y=292)

calender_dateto = DateEntry(master=l1, width=8, background= "#000000", bd=5)
calender_dateto.place(x=171, y=319)


label_weekday=ctk.CTkLabel(master=l1, text="Week Day", bg_color='#ACACAC', font=('Century Gothic',12))
label_weekday.place(x=296, y=292)

def combobox_weekday(choice):
    print("combobox dropdown clicked:", choice)

combobox_week_day = ctk.CTkComboBox(master=l1,
                                    values=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday","Sunday"],
                                     command=combobox_weekday,
                                     bg_color='#ACACAC',
                                     width=86)
combobox_week_day.place(x=283, y=319)
combobox_week_day.set("Monday")  # set initial value

label_duration=ctk.CTkLabel(master=l1, text="Duration", bg_color='#ACACAC', font=('Century Gothic',12))
label_duration.place(x=190, y=480)

entry_duration_monday=ctk.CTkEntry(master=l1, width=250, placeholder_text="Enter in this format: 8,0 or 8:00.", bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_duration_monday.place(x=90, y=500)

entry_duration_tuesday=ctk.CTkEntry(master=l1, width=250, placeholder_text="Enter in this format: 8,0 or 8:00.", bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_duration_tuesday.place(x=470, y=200)

entry_duration_wednesday=ctk.CTkEntry(master=l1, width=250, placeholder_text="Enter in this format: 8,0 or 8:00.", bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_duration_wednesday.place(x=870, y=200)

entry_duration_thursday=ctk.CTkEntry(master=l1, width=250, placeholder_text="Enter in this format: 8,0 or 8:00.", bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_duration_thursday.place(x=470, y=385)

entry_duration_friday=ctk.CTkEntry(master=l1, width=250, placeholder_text="Enter in this format: 8,0 or 8:00.", bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_duration_friday.place(x=870, y=385)

label_content=ctk.CTkLabel(master=l1, text="Content", bg_color='#ACACAC', font=('Century Gothic',12))
label_content.place(x=190, y=350)

entry_content_monday=ctk.CTkTextbox(master=l1, height=100, width=312, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_content_monday.place(x=59, y=374)

entry_content_tuesday=ctk.CTkTextbox(master=l1, height=100, width=312, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_content_tuesday.place(x=450, y=75)

entry_content_wendesday=ctk.CTkTextbox(master=l1, height=100, width=312, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_content_wendesday.place(x=850, y=75)

entry_content_thursday=ctk.CTkTextbox(master=l1, height=100, width=312, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_content_thursday.place(x=450, y=260)

entry_content_friday=ctk.CTkTextbox(master=l1, height=100, width=312, bg_color='#ACACAC', corner_radius=20, font=('Century Gothic',12))
entry_content_friday.place(x=850, y=260)



button= ctk.CTkButton(master=l1, text="Save", command=enter_data, width=170, height=40, compound="left", bg_color='#A3A3A3', fg_color='white', text_color='#7c7c7c', hover_color='#000000')
button.place(x=90, y=580)

button= ctk.CTkButton(master=l1, text="Export to Word", command=excel_to_word, width=170, height=40, compound="left", bg_color='#A3A3A3', fg_color='white', text_color='#7c7c7c', hover_color='#000000')
button.place(x=300, y=580)

button= ctk.CTkButton(master=l1, text="Export to PDF", command=word_to_pdf, width=170, height=40, compound="left", bg_color='#A3A3A3', fg_color='white', text_color='#7c7c7c', hover_color='#000000')
button.place(x=500, y=580)

button= ctk.CTkButton(master=l1, text="Load sheet", command=load_file, width=170, height=40, compound="left", bg_color='#A3A3A3', fg_color='white', text_color='#7c7c7c', hover_color='#000000')
button.place(x=700, y=580)

button= ctk.CTkButton(master=l1, text="Clear", command=clear, width=170, height=40, compound="left", bg_color='#A3A3A3', fg_color='white', text_color='#7c7c7c', hover_color='#000000')
button.place(x=900, y=580)

button= ctk.CTkButton(master=l1, text="Close", command=close, width=80, height=40, compound="left", bg_color='#A3A3A3', fg_color='white', text_color='#000000', hover_color='#c01f1f')
button.place(x=1080, y=580)

app.mainloop()


# user_info_frame =tkinter.LabelFrame(app, text="User Information")
# user_info_frame.grid(row= 0, column=0, padx=20, pady=10)

# first_name_label = tkinter.Label(user_info_frame, text="First Name")
# first_name_label.grid(row=0, column=0)
# last_name_label = tkinter.Label(user_info_frame, text="Last Name")
# last_name_label.grid(row=0, column=1)

# first_name_entry = tkinter.Entry(user_info_frame)
# last_name_entry = tkinter.Entry(user_info_frame)
# first_name_entry.grid(row=1, column=0)
# last_name_entry.grid(row=1, column=1)

# title_label = tkinter.Label(user_info_frame, text="Title")
# title_combobox = ttk.Combobox(user_info_frame, values=["", "Mr.", "Ms.", "Dr."])
# title_label.grid(row=0, column=2)
# title_combobox.grid(row=1, column=2)

# age_label = tkinter.Label(user_info_frame, text="Age")
# age_spinbox = tkinter.Spinbox(user_info_frame, from_=18, to=110)
# age_label.grid(row=2, column=0)
# age_spinbox.grid(row=3, column=0)

# nationality_label = tkinter.Label(user_info_frame, text="Nationality")
# nationality_combobox = ttk.Combobox(user_info_frame, values=["Africa", "Antarctica", "Asia", "Europe", "North America", "Oceania", "South America"])
# nationality_label.grid(row=2, column=1)
# nationality_combobox.grid(row=3, column=1)

# for widget in user_info_frame.winfo_children():
#     widget.grid_configure(padx=10, pady=5)

# # Saving Course Info
# courses_frame = tkinter.LabelFrame(app)
# courses_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)

# registered_label = tkinter.Label(courses_frame, text="Registration Status")

# reg_status_var = tkinter.StringVar(value="Not Registered")
# registered_check = tkinter.Checkbutton(courses_frame, text="Currently Registered",
#                                        variable=reg_status_var, onvalue="Registered", offvalue="Not registered")

# registered_label.grid(row=0, column=0)
# registered_check.grid(row=1, column=0)

# numcourses_label = tkinter.Label(courses_frame, text= "# Completed Courses")
# numcourses_spinbox = tkinter.Spinbox(courses_frame, from_=0, to='infinity')
# numcourses_label.grid(row=0, column=1)
# numcourses_spinbox.grid(row=1, column=1)

# numsemesters_label = tkinter.Label(courses_frame, text="# Semesters")
# numsemesters_spinbox = tkinter.Spinbox(courses_frame, from_=0, to="infinity")
# numsemesters_label.grid(row=0, column=2)
# numsemesters_spinbox.grid(row=1, column=2)

# for widget in courses_frame.winfo_children():
#     widget.grid_configure(padx=10, pady=5)

# # Accept terms
# terms_frame = tkinter.LabelFrame(app, text="Terms & Conditions")
# terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

# accept_var = tkinter.StringVar(value="Not Accepted")
# terms_check = tkinter.Checkbutton(terms_frame, text= "I accept the terms and conditions.",
#                                   variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
# terms_check.grid(row=0, column=0)

# # Button
# button = tkinter.Button(app, text="Enter data", command= enter_data)
# button.grid(row=3, column=0, sticky="news", padx=20, pady=10)
 
# app.mainloop()